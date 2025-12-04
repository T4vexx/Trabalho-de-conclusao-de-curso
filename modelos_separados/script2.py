
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Carregar o Excel existente
xlsx_path = 'resultados_todos_modelos_consolidado.xlsx'
df_otimizados = pd.read_csv('modelos_otimizados_resumo.csv')

# Carregar workbook
wb = load_workbook(xlsx_path)

print("=" * 80)
print("ABAS EXISTENTES NO EXCEL:")
print("=" * 80)
for sheet_name in wb.sheetnames:
    print(f"  • {sheet_name}")

# Criar nova aba "Modelos Otimizados"
nova_aba_nome = "Modelos Otimizados (Grid Search)"
if nova_aba_nome in wb.sheetnames:
    del wb[nova_aba_nome]
    print(f"\n⚠️  Aba '{nova_aba_nome}' já existia - substituída")

ws = wb.create_sheet(nova_aba_nome)

# Estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
highlight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
normal_font = Font(name='Calibri', size=10)
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Título
ws.merge_cells('A1:I1')
ws['A1'] = "📊 RESULTADOS DOS MODELOS OTIMIZADOS COM GRID/RANDOM SEARCH"
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="1F4E78")
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Subtítulo
ws.merge_cells('A2:I2')
ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
ws['A2'].font = Font(name='Calibri', size=9, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
row_num = 4
headers = ['Modelo Embedding', 'Dataset', 'Classificador', 'Método', 
           'F1-Score', 'Accuracy', 'CV Score', 'GPU', 'Melhores Parâmetros']

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_thin

# Adicionar dados
row_num = 5
max_f1 = df_otimizados['Test_F1'].max()

for idx, row_data in df_otimizados.iterrows():
    # Modelo Embedding
    cell = ws.cell(row=row_num, column=1)
    cell.value = row_data['Embedding_Model']
    cell.font = Font(name='Calibri', size=10, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border_thin
    
    # Dataset
    cell = ws.cell(row=row_num, column=2)
    cell.value = row_data['Dataset']
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border_thin
    
    # Classificador
    cell = ws.cell(row=row_num, column=3)
    cell.value = row_data['Classifier']
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border_thin
    
    # Método
    cell = ws.cell(row=row_num, column=4)
    cell.value = row_data['Method']
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
    
    # F1-Score
    cell = ws.cell(row=row_num, column=5)
    cell.value = row_data['Test_F1']
    cell.number_format = '0.0000'
    cell.font = Font(name='Calibri', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
    # Destacar melhor resultado
    if row_data['Test_F1'] == max_f1:
        cell.fill = best_fill
    
    # Accuracy
    cell = ws.cell(row=row_num, column=6)
    cell.value = row_data['Test_Accuracy']
    cell.number_format = '0.0000'
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
    
    # CV Score
    cell = ws.cell(row=row_num, column=7)
    if pd.notna(row_data['CV_Score']):
        cell.value = row_data['CV_Score']
        cell.number_format = '0.0000'
    else:
        cell.value = 'N/A'
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
    
    # GPU
    cell = ws.cell(row=row_num, column=8)
    if row_data['Used_GPU']:
        cell.value = '✅'
    else:
        cell.value = '❌'
    cell.font = Font(name='Calibri', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
    
    # Parâmetros
    cell = ws.cell(row=row_num, column=9)
    params_str = str(row_data['Best_Params'])
    # Simplificar params
    if len(params_str) > 100:
        params_str = params_str[:100] + "..."
    cell.value = params_str
    cell.font = Font(name='Calibri', size=9)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border_thin
    
    row_num += 1

# Adicionar resumo estatístico
row_num += 2
ws.merge_cells(f'A{row_num}:I{row_num}')
ws[f'A{row_num}'] = "📈 RESUMO ESTATÍSTICO POR MODELO DE EMBEDDING"
ws[f'A{row_num}'].font = Font(name='Calibri', size=12, bold=True, color="1F4E78")
ws[f'A{row_num}'].alignment = Alignment(horizontal='center')

row_num += 1
resumo_headers = ['Modelo', 'F1 Médio', 'F1 Máximo', 'F1 Mínimo', 
                  'Acc Médio', 'Acc Máximo', 'Acc Mínimo', 'Nº Testes']

for col_num, header in enumerate(resumo_headers, 1):
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = header
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin

row_num += 1

# Calcular estatísticas
for modelo in df_otimizados['Embedding_Model'].unique():
    df_modelo = df_otimizados[df_otimizados['Embedding_Model'] == modelo]
    
    stats = [
        modelo,
        df_modelo['Test_F1'].mean(),
        df_modelo['Test_F1'].max(),
        df_modelo['Test_F1'].min(),
        df_modelo['Test_Accuracy'].mean(),
        df_modelo['Test_Accuracy'].max(),
        df_modelo['Test_Accuracy'].min(),
        len(df_modelo)
    ]
    
    for col_num, value in enumerate(stats, 1):
        cell = ws.cell(row=row_num, column=col_num)
        if col_num == 1:
            cell.value = value
            cell.font = Font(name='Calibri', size=10, bold=True)
        elif col_num == 8:
            cell.value = value
            cell.font = normal_font
        else:
            cell.value = value
            cell.number_format = '0.0000'
            cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    row_num += 1

# Ajustar larguras
ws.column_dimensions['A'].width = 20  # Modelo
ws.column_dimensions['B'].width = 25  # Dataset
ws.column_dimensions['C'].width = 20  # Classificador
ws.column_dimensions['D'].width = 15  # Método
ws.column_dimensions['E'].width = 12  # F1
ws.column_dimensions['F'].width = 12  # Acc
ws.column_dimensions['G'].width = 12  # CV
ws.column_dimensions['H'].width = 8   # GPU
ws.column_dimensions['I'].width = 60  # Parâmetros

# Ajustar altura das linhas de dados
for row in range(5, row_num):
    ws.row_dimensions[row].height = 30

# Salvar
output_path = 'resultados_todos_modelos_consolidado_v2.xlsx'
wb.save(output_path)

print("\n" + "=" * 80)
print("✅ EXCEL ATUALIZADO COM SUCESSO!")
print("=" * 80)
print(f"📁 Arquivo: {output_path}")
print(f"📊 Nova aba: '{nova_aba_nome}'")
print(f"📈 Total de resultados: {len(df_otimizados)}")
print(f"🏆 Melhor F1-Score: {max_f1:.4f}")
print(f"\n🔍 Abas no arquivo:")
for sheet_name in wb.sheetnames:
    print(f"  • {sheet_name}")
